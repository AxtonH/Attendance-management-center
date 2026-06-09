"""Render an `ExportTable` to an .xlsx workbook (openpyxl).

The only place that knows the Excel format. Takes the renderer-agnostic
`ExportTable` from `exports.service` and returns the file as raw bytes, ready
to stream as a download. No attendance logic here — just layout.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.features.exports.service import ExportTable

# Prezlab dark header band, white text — echoes the app's bg-bg-dark bars.
_HEADER_FILL = PatternFill("solid", fgColor="1F2430")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
# Per-employee section banner in the grouped (weekly/monthly) layout.
_SECTION_FILL = PatternFill("solid", fgColor="EEF1F5")
_SECTION_FONT = Font(bold=True, color="1F2430")
_TITLE_FONT = Font(bold=True, size=14)
_SUBTITLE_FONT = Font(size=10, color="6B7280")

# Column widths (chars) tuned to the employee table: code, name, two times,
# worked/status. Name is widest; times are narrow.
_COLUMN_WIDTHS = [18, 26, 12, 12, 18]


def render_xlsx(table: ExportTable) -> bytes:
    """Build the workbook and return it as bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    n_cols = len(table.columns)
    last_col_letter = get_column_letter(n_cols)

    # Title + subtitle band (rows 1–2), merged across the table width.
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = table.title
    title_cell.font = _TITLE_FONT

    ws.merge_cells(f"A2:{last_col_letter}2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = table.subtitle
    subtitle_cell.font = _SUBTITLE_FONT

    header_row_idx = 4  # leave row 3 blank as a spacer
    for col_idx, name in enumerate(table.columns, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    for offset, row_values in enumerate(table.rows):
        excel_row = header_row_idx + 1 + offset
        is_section = offset in table.section_rows
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            if is_section:
                cell.fill = _SECTION_FILL
                cell.font = _SECTION_FONT

    _apply_widths(ws, n_cols)
    # Freeze the header so it stays visible while scrolling a long roster.
    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _apply_widths(ws: Worksheet, n_cols: int) -> None:
    for col_idx in range(1, n_cols + 1):
        width = (
            _COLUMN_WIDTHS[col_idx - 1]
            if col_idx - 1 < len(_COLUMN_WIDTHS)
            else 16
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = width
